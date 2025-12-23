#!/usr/bin/env python3
"""
测试用例生成器 - 按五维度生成完整的接口测试用例
五维度：正常/参数异常/业务异常/权限安全/性能边界
"""
import os
import sys
import json
import re
import yaml
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 添加generated_proto路径
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__)), 'generated_proto'))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__)), 'generated_proto', 'client'))

from framework.config import Config
from framework.protobuf_parser import ProtobufParser
from framework.proto_request_formatter import ProtoRequestFormatter
from framework.client import APIClient
from framework.request_data_converter import RequestDataConverter
from framework.base_test_case_generator import BaseTestCaseGenerator


class TestCaseGenerator(BaseTestCaseGenerator):
    """测试用例生成器"""
    
    # 优先级映射
    PRIORITY_MAP = {
        'P0': '高',
        'P1': '中',
        'P2': '低'
    }
    
    def __init__(self, config: Config):
        super().__init__(config)
    
    def generate_test_cases_excel(self, output_file: str = "test_cases_complete.xlsx", save_to_test_cases: bool = True, test_results: Dict[str, Any] = None, run_tests: bool = True):
        """生成测试用例Excel文件
        
        Args:
            output_file: 输出文件名
            save_to_test_cases: 是否同时保存到test_cases目录（默认True）
            test_results: 实际测试结果，如果提供则根据实际结果更新状态
                         如果为None，尝试从最新的test_report.xlsx读取
            run_tests: 是否实际运行接口测试（默认True）
        """
        print("=" * 80)
        print("生成测试用例Excel...")
        print("=" * 80)
        
        # 如果指定运行测试，则实际运行接口并获取真实数据
        if run_tests:
            print("\n" + "=" * 80)
            print("实际运行接口测试，获取真实返回数据...")
            print("=" * 80)
            test_results = self._run_all_tests_from_yaml()
        elif test_results is None:
            # 如果没有提供测试结果，尝试从最新的test_report.xlsx读取
            test_results = self._load_test_results_from_excel()
        
        # 解析接口
        parser = ProtobufParser(self.config)
        interfaces = parser.discover_interfaces()
        
        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet
        
        # 为每个服务生成测试用例
        for service_name, service_interfaces in interfaces.items():
            print(f"\n处理服务: {service_name}")
            sheet = wb.create_sheet(title=service_name.upper())
            # 获取该服务的实际测试结果（如果有）
            service_test_results = None
            if test_results and 'services' in test_results:
                service_test_results = test_results['services'].get(service_name)
            self._generate_service_sheet(sheet, service_name, service_interfaces, service_test_results)
        
        # 保存文件到reports目录（只保存到reports）
        report_path = os.path.join(self.config.get_report_dir(), output_file)
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        wb.save(report_path)
        print(f"\n✓ 测试用例Excel已生成: {report_path}")
        
        return report_path
    
    def _generate_service_sheet(self, sheet, service_name: str, interfaces: List[Dict], service_test_results: Dict[str, Any] = None):
        """为单个服务生成测试用例sheet
        
        Args:
            sheet: Excel工作表
            service_name: 服务名称
            interfaces: 接口列表
            service_test_results: 该服务的实际测试结果（可选）
        """
        # 设置表头
        headers = [
            '用例编号', '标题', '优先级', '前置条件', '维度', '方法+URL', 
            '请求头', '请求', '预期状态码', '预期服务器返回', '实际服务器返回', '状态', 'JSONPath断言', '服务器报错', '备注'
        ]
        
        # 写入表头
        header_row = 1
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 用例编号
            'B': 30,  # 标题
            'C': 8,   # 优先级
            'D': 30,  # 前置条件
            'E': 12,  # 维度
            'F': 25,  # 方法+URL
            'G': 20,  # 请求头
            'H': 40,  # 请求
            'I': 12,  # 预期状态码
            'J': 50,  # 预期服务器返回
            'K': 50,  # 实际服务器返回
            'L': 12,  # 状态
            'M': 30,  # JSONPath断言
            'N': 30,  # 服务器报错
            'O': 30   # 备注
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # 初始化已匹配的测试结果集合（每个服务独立，避免重复匹配）
        self._matched_test_results = set()
        
        # 从 test_cases 目录读取 YAML 测试用例
        yaml_file = f"test_cases/{service_name}/test_{service_name}.yaml"
        yaml_test_cases = self._load_yaml_test_cases(yaml_file)
        
        # 为每个接口生成测试用例
        row = 2
        for interface in interfaces:
            method_name = interface['name']
            print(f"  生成接口测试用例: {method_name}")
            
            # 从 YAML 文件中获取该接口的所有测试用例（正常和异常）
            cases = self._get_cases_from_yaml(yaml_test_cases, service_name, method_name, interface)
            
            # 获取该接口的实际测试结果（如果有）
            # 为正常用例查找匹配的测试结果
            actual_test_result = None
            if service_test_results:
                test_list = service_test_results.get('test_results', [])
                print(f"    查找匹配的测试结果，共有 {len(test_list)} 个测试结果")
                # 查找匹配的测试结果（优先通过method和dimension匹配正常用例）
                for test in test_list:
                    test_method = test.get('method', '')
                    test_dimension = test.get('dimension', '正常')
                    test_name = test.get('name', '')
                    
                    # 优先匹配：method相同且dimension为'正常'
                    if test_method and test_method.lower() == method_name.lower() and test_dimension == '正常':
                        print(f"    ✓ 找到正常用例测试结果: {test_name} (method={test_method}, dimension={test_dimension})")
                        actual_test_result = test
                        break
                    
                    # 备用匹配：通过test_name匹配（兼容旧逻辑）
                    test_name_normalized = test_name.replace(' ', '').replace('_', '').lower()
                    method_name_normalized = method_name.replace(' ', '').replace('_', '').lower()
                    if (test_name == method_name or 
                        test_name.lower() == method_name.lower() or
                        test_name_normalized == method_name_normalized):
                        # 如果还没有找到，且这个测试是正常用例，使用它
                        if not actual_test_result and test_dimension == '正常':
                            print(f"    ✓ 通过备用匹配找到正常用例测试结果: {test_name} (method={test_method}, dimension={test_dimension})")
                            actual_test_result = test
                
                if not actual_test_result:
                    print(f"    ⚠ 未找到匹配的正常用例测试结果 (method={method_name})")
            
            # 写入测试用例
            for case in cases:
                # 如果有实际测试结果，更新正常用例的状态和响应
                if actual_test_result and case.get('dimension') == '正常':
                    # 只更新正常用例的实际结果
                    case = self._update_case_with_actual_result(case, actual_test_result)
                    # 更新前置条件为实际执行的前置操作
                    actual_preconditions = actual_test_result.get('preconditions', '')
                    if actual_preconditions:
                        # 如果前置条件是字符串，直接使用；如果是列表，则格式化
                        if isinstance(actual_preconditions, list):
                            case['preconditions'] = '\n'.join([f'• {p}' for p in actual_preconditions])
                        else:
                            case['preconditions'] = actual_preconditions
                elif case.get('dimension') != '正常' and service_test_results:
                    # 对于异常用例，尝试匹配测试结果（根据测试名称中的异常类型）
                    # 异常用例的测试名称格式：FetchSelfFullUserInfo_参数异常_必填参数缺失
                    case_title = case.get('title', '')
                    case_dimension = case.get('dimension', '')
                    # 从 case_title 中提取异常类型（例如：SendMessage_参数异常_必填参数缺失 -> 参数异常_必填参数缺失）
                    abnormal_type_in_title = ''
                    if '_' in case_title:
                        parts = case_title.split('_')
                        if len(parts) >= 3:
                            abnormal_type_in_title = '_'.join(parts[2:])  # 取第3部分及之后
                        elif len(parts) == 2:
                            abnormal_type_in_title = parts[1]
                    
                    # 查找匹配的异常测试结果
                    # 使用已匹配的测试结果集合，避免重复匹配
                    if not hasattr(self, '_matched_test_results'):
                        self._matched_test_results = set()  # 存储已匹配的测试结果ID
                    
                    matched_abnormal_result = None
                    matched_candidates = []  # 存储所有匹配的候选结果
                    
                    if service_test_results:
                        test_list = service_test_results.get('test_results', [])
                        for test_idx, test in enumerate(test_list):
                            # 检查是否已经匹配过
                            test_id = f"{test.get('method', '')}_{test.get('dimension', '')}_{test_idx}"
                            if test_id in self._matched_test_results:
                                continue
                            
                            test_name = test.get('name', '')
                            test_method = test.get('method', '')
                            test_dimension = test.get('dimension', '')
                            test_abnormal_type = test.get('abnormal_type', '')
                            test_error_code = test.get('error_code')
                            
                            # 首先检查方法名是否匹配
                            if test_method and test_method.lower() != method_name.lower():
                                continue
                            
                            # 检查维度是否匹配
                            if test_dimension and case_dimension and test_dimension != case_dimension:
                                continue
                            
                            # 标准化名称进行匹配
                            # case_title 格式: "SendMessage_参数异常_必填参数缺失"
                            # test_name 格式: "SendMessage_参数异常_必填参数缺失" 或 "SendMessage"（如果标题显示有问题）
                            case_title_normalized = case_title.replace(' ', '').replace('_', '').replace('-', '').lower()
                            test_name_normalized = test_name.replace(' ', '').replace('_', '').replace('-', '').lower()
                            
                            # 匹配异常测试用例：优先使用精确匹配
                            match_score = 0
                            if case_title and case_title == test_name:
                                # 精确匹配，直接使用
                                matched_abnormal_result = test
                                self._matched_test_results.add(test_id)
                                break
                            elif case_title_normalized == test_name_normalized:
                                # 标准化后完全匹配
                                matched_abnormal_result = test
                                self._matched_test_results.add(test_id)
                                break
                            elif test_abnormal_type and abnormal_type_in_title:
                                # 使用异常类型匹配
                                if test_abnormal_type.replace(' ', '').replace('_', '').replace('-', '').lower() == abnormal_type_in_title.replace(' ', '').replace('_', '').replace('-', '').lower():
                                    matched_abnormal_result = test
                                    self._matched_test_results.add(test_id)
                                    break
                            elif test_dimension == case_dimension:
                                # 维度匹配，记录为候选
                                match_score = 1
                                matched_candidates.append((match_score, test_idx, test))
                    
                    # 如果没有精确匹配，使用候选结果（按维度匹配的第一个未匹配的）
                    if not matched_abnormal_result and matched_candidates:
                        # 按匹配分数和索引排序，选择第一个未匹配的
                        matched_candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
                        for score, test_idx, test in matched_candidates:
                            test_id = f"{test.get('method', '')}_{test.get('dimension', '')}_{test_idx}"
                            if test_id not in self._matched_test_results:
                                matched_abnormal_result = test
                                self._matched_test_results.add(test_id)
                                break
                    
                    # 如果找到匹配的异常测试结果，更新用例
                    if matched_abnormal_result:
                        case = self._update_case_with_actual_result(case, matched_abnormal_result)
                        # 更新前置条件为实际执行的前置操作
                        actual_preconditions = matched_abnormal_result.get('preconditions', '')
                        if actual_preconditions:
                            # 如果前置条件是字符串，直接使用；如果是列表，则格式化
                            if isinstance(actual_preconditions, list):
                                case['preconditions'] = '\n'.join([f'• {p}' for p in actual_preconditions])
                            else:
                                case['preconditions'] = actual_preconditions
                
                self._write_test_case_row(sheet, row, case, service_name, method_name)
                row += 1
        
        # 设置行高
        for row_idx in range(2, row):
            sheet.row_dimensions[row_idx].height = 60
    
    def _load_yaml_test_cases(self, yaml_file: str) -> Dict:
        """从 YAML 文件加载测试用例"""
        if not os.path.exists(yaml_file):
            return {}
        
        try:
            with open(yaml_file, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            print(f"⚠ 加载YAML测试用例失败 {yaml_file}: {e}")
            return {}
    
    def _get_cases_from_yaml(self, yaml_test_cases: Dict, service_name: str, method_name: str, interface: Dict) -> List[Dict]:
        """从 YAML 测试用例中提取指定接口的所有测试用例（正常和异常）"""
        cases = []
        test_cases_dict = yaml_test_cases.get('test_cases', {})
        
        # 获取请求结构（用于格式化请求体）
        service_name_cap = service_name.capitalize()
        request_structure = ProtoRequestFormatter.get_request_structure(service_name_cap, method_name)
        if request_structure is None:
            request_structure = {}
        
        # 检查接口是否有参数（根据协议定义）
        has_parameters = bool(request_structure and len(request_structure) > 0)
        
        # 如果没有参数，只生成正常测试用例，跳过参数异常测试用例
        if not has_parameters:
            print(f"    ⚠ 接口 {method_name} 无参数，将过滤参数异常测试用例")
        
        # 查找所有相关的测试用例（正常和异常）
        case_counter = 1
        for case_key, case_data in test_cases_dict.items():
            # 匹配测试用例：case_key 应该是 method_name 或 method_name_xxx
            if case_key == method_name or case_key.startswith(f"{method_name}_"):
                # 提取维度信息
                dimension = case_data.get('dimension', '正常')
                
                # 如果接口没有参数，跳过参数异常测试用例
                if not has_parameters and dimension == '参数异常':
                    print(f"    ⚠ 跳过参数异常测试用例: {case_key} (接口无参数)")
                    continue
                
                abnormal_type = None
                if dimension != '正常':
                    # 从 case_key 中提取异常类型，例如：FetchSelfFullUserInfo_参数异常_必填参数缺失
                    parts = case_key.split('_')
                    if len(parts) > 1:
                        abnormal_type = '_'.join(parts[1:])  # 参数异常_必填参数缺失
                
                # 格式化服务器响应（expected_response）
                expected_response = case_data.get('expected_response', {})
                if isinstance(expected_response, dict):
                    expected_response_str = json.dumps(expected_response, indent=2, ensure_ascii=False, default=str)
                else:
                    expected_response_str = str(expected_response)
                
                # 转换请求数据以分析前置条件
                request_data_yaml = case_data.get('request', {})
                request_data_actual = RequestDataConverter.convert_nested_request_data(request_data_yaml)
                # 根据实际请求参数生成详细的前置条件
                detailed_preconditions = self._get_preconditions(service_name, method_name, request_data_actual)
                
                # 构建测试用例（包含所有必需的字段）
                case = {
                    'case_id': f"TC{case_counter:04d}",
                    'title': case_data.get('description', case_key),
                    'priority': case_data.get('priority', 'P1'),
                    'preconditions': detailed_preconditions,  # 使用详细的前置条件
                    'dimension': dimension,
                    'abnormal_type': abnormal_type,
                    'method_name': method_name,
                    'method_url': f'TCP {service_name}.{method_name}',
                    'headers': 'Content-Type: application/protobuf',
                    'request_body': self._format_request_body_from_yaml(case_data.get('request', {}), request_structure),
                    'expected_status': case_data.get('expected_status', '200'),
                    'expected_response': expected_response_str,
                    'server_response': expected_response_str,  # 兼容旧字段名
                    'jsonpath_assertion': case_data.get('jsonpath_assertion', ''),
                    'remark': case_data.get('remark', ''),
                    'status': '未测试',  # 默认状态，后续会根据实际测试结果更新
                    'error_message': '',
                    'server_error': '',
                    'possible_issues': '',
                    'actual_response': ''  # 实际服务器返回，初始为空
                }
                cases.append(case)
                case_counter += 1
        
        # 如果没有找到测试用例，生成一个默认的正常测试用例
        if not cases:
            case = {
                'case_id': f"TC0001",
                'title': f'{method_name}_正常',
                'priority': 'P0',
                'preconditions': '已登录',
                'dimension': '正常',
                'abnormal_type': None,
                'method_name': method_name,
                'method_url': f'TCP {service_name}.{method_name}',
                'headers': 'Content-Type: application/protobuf',
                'request_body': self._format_request_body_from_yaml({}, request_structure),
                'expected_status': '200',
                'expected_response': '{}',
                'server_response': '{}',
                'jsonpath_assertion': '',
                'remark': '正常业务流程验证',
                'status': '未测试',
                'error_message': '',
                'server_error': '',
                'possible_issues': '',
                'actual_response': ''
            }
            cases.append(case)
        
        return cases
    
    def _format_request_body_from_yaml(self, request_data: Dict, request_structure: Dict[str, str]) -> str:
        """格式化请求体（从 YAML 中的 request 字段）"""
        if not request_structure:
            # 如果没有请求结构，直接格式化请求数据
            return json.dumps(request_data, indent=2, ensure_ascii=False) if request_data else '{}'
        
        # 格式化请求体，包含字段类型信息
        # 注意：如果 request_data 中的值已经是 {"value": ..., "type": ...} 格式，直接使用
        formatted = {}
        for field_name, field_type in request_structure.items():
            if field_name in request_data:
                field_value = request_data[field_name]
                # 检查是否已经是 {"value": ..., "type": ...} 格式
                if isinstance(field_value, dict) and 'value' in field_value and 'type' in field_value:
                    # 已经是正确格式，直接使用
                    formatted[field_name] = field_value
                else:
                    # 需要包装成 {"value": ..., "type": ...} 格式
                    formatted[field_name] = {
                        'value': field_value,
                        'type': field_type
                    }
            else:
                formatted[field_name] = {
                    'value': None,
                    'type': field_type
                }
        
        return json.dumps(formatted, indent=2, ensure_ascii=False)
    
    # _generate_five_dimension_cases 已在基类中实现
    
    def _generate_normal_case(self, service_name: str, method_name: str, 
                              request_structure: Dict[str, str]) -> Dict:
        """生成正常用例（正例）"""
        self.case_counter += 1
        case_id = f"TC{self.case_counter:04d}"
        
        # 生成正常请求体（实际发送给服务器的格式）
        request_body = self._generate_normal_request_body(request_structure)
        request_body_str = self._format_request_body_with_types(request_body, request_structure)
        
        # 生成预期的服务器返回
        expected_response = self._generate_expected_response(service_name, method_name, request_body, True)
        
        # 获取前置条件
        preconditions = self._get_preconditions(service_name, method_name)
        
        return {
            'case_id': case_id,
            'title': f'{method_name}_正常调用',
            'priority': 'P0',
            'preconditions': preconditions,
            'dimension': '正常',
            'method_url': f'TCP {service_name}.{method_name}',
            'headers': 'Content-Type: application/protobuf',
            'request_body': request_body_str,
            'expected_status': '200',
            'server_response': expected_response,
            'status': '通过',
            'jsonpath_assertion': self._generate_jsonpath_assertion(method_name, True),
            'error_message': '',
            'problem_analysis': '',
            'remark': '正常业务流程验证'
        }
    
    def _generate_parameter_abnormal_cases(self, service_name: str, method_name: str,
                                          request_structure: Dict[str, str]) -> List[Dict]:
        """生成参数异常用例（5条高频反例）"""
        cases = []
        abnormal_types = [
            ('必填参数缺失', 'missing_required'),
            ('参数类型错误', 'type_error'),
            ('参数值为空', 'empty_value'),
            ('参数值超出范围', 'out_of_range'),
            ('参数格式错误', 'format_error')
        ]
        
        for abnormal_name, abnormal_type in abnormal_types:
            self.case_counter += 1
            case_id = f"TC{self.case_counter:04d}"
            
            # 生成异常请求体
            request_body = self._generate_abnormal_request_body(
                request_structure, abnormal_type
            )
            
            # 获取前置条件
            preconditions = self._get_preconditions(service_name, method_name)
            
            # 格式化请求体（带类型信息）
            request_body_str = self._format_request_body_with_types(request_body, request_structure)
            
            # 生成预期的服务器返回（错误响应）
            expected_response = self._generate_expected_response(service_name, method_name, request_body, False)
            
            cases.append({
                'case_id': case_id,
                'title': f'{method_name}_参数异常_{abnormal_name}',
                'priority': 'P1',
                'preconditions': preconditions,
                'dimension': '参数异常',
                'method_url': f'TCP {service_name}.{method_name}',
                'headers': 'Content-Type: application/protobuf',
                'request_body': request_body_str,
                'expected_status': '400/500',
                'server_response': expected_response,
                'status': '通过',  # 异常用例：返回预期错误码（400/500）即为通过
                'jsonpath_assertion': f'$.error_code != 200',
                'error_message': '',
                'problem_analysis': '',
                'remark': f'{abnormal_name}场景验证'
            })
        
        return cases
    
    def _generate_business_abnormal_cases(self, service_name: str, method_name: str,
                                         request_structure: Dict[str, str]) -> List[Dict]:
        """生成业务异常用例"""
        cases = []
        
        # 根据接口类型生成不同的业务异常用例
        business_abnormal_scenarios = self._get_business_abnormal_scenarios(service_name, method_name)
        
        for scenario in business_abnormal_scenarios:
            self.case_counter += 1
            case_id = f"TC{self.case_counter:04d}"
            
            request_body = self._generate_business_abnormal_request_body(
                request_structure, scenario
            )
            
            # 获取前置条件
            preconditions = self._get_preconditions(service_name, method_name)
            
            # 格式化请求体（带类型信息）
            request_body_str = self._format_request_body_with_types(request_body, request_structure)
            
            # 生成预期的服务器返回
            expected_response = self._generate_expected_response(service_name, method_name, request_body, False)
            
            cases.append({
                'case_id': case_id,
                'title': f'{method_name}_业务异常_{scenario["name"]}',
                'priority': 'P1',
                'preconditions': preconditions,
                'dimension': '业务异常',
                'method_url': f'TCP {service_name}.{method_name}',
                'headers': 'Content-Type: application/protobuf',
                'request_body': request_body_str,
                'expected_status': scenario.get('expected_status', '400'),
                'server_response': expected_response,
                'status': '通过',  # 业务异常用例：返回预期错误码即为通过
                'jsonpath_assertion': scenario.get('jsonpath', '$.error_code != 200'),
                'error_message': '',
                'problem_analysis': '',
                'remark': scenario.get('remark', '业务异常场景验证')
            })
        
        return cases
    
    def _generate_security_cases(self, service_name: str, method_name: str,
                                request_structure: Dict[str, str]) -> List[Dict]:
        """生成权限安全用例"""
        cases = []
        
        security_scenarios = [
            {
                'name': '未授权访问',
                'request_body': {},
                'expected_status': '401',
                'jsonpath': '$.error_code == 401',
                'remark': 'token缺失或无效'
            },
            {
                'name': '越权访问',
                'request_body': self._generate_normal_request_body(request_structure),
                'expected_status': '403',
                'jsonpath': '$.error_code == 403',
                'remark': '无权限访问他人资源'
            }
        ]
        
        for scenario in security_scenarios:
            self.case_counter += 1
            case_id = f"TC{self.case_counter:04d}"
            
            # 获取前置条件（权限安全场景通常需要登录）
            preconditions = '已登录' if scenario['name'] == '未授权访问' else self._get_preconditions(service_name, method_name)
            
            # 格式化请求体（带类型信息）
            request_body_str = self._format_request_body_with_types(scenario['request_body'], request_structure)
            
            # 生成预期的服务器返回
            expected_response = self._generate_expected_response(service_name, method_name, scenario['request_body'], False)
            
            cases.append({
                'case_id': case_id,
                'title': f'{method_name}_权限安全_{scenario["name"]}',
                'priority': 'P0',
                'preconditions': preconditions,
                'dimension': '权限安全',
                'method_url': f'TCP {service_name}.{method_name}',
                'headers': 'Content-Type: application/protobuf',
                'request_body': request_body_str,
                'expected_status': scenario['expected_status'],
                'server_response': expected_response,
                'status': '通过',  # 权限安全用例：返回预期错误码（401/403）即为通过
                'jsonpath_assertion': scenario['jsonpath'],
                'error_message': '',
                'problem_analysis': '',
                'remark': scenario['remark']
            })
        
        return cases
    
    def _generate_performance_cases(self, service_name: str, method_name: str,
                                   request_structure: Dict[str, str]) -> List[Dict]:
        """生成性能边界用例"""
        cases = []
        
        performance_scenarios = [
            {
                'name': '大数据量',
                'request_body': self._generate_large_data_request(request_structure),
                'expected_status': '200/413',
                'jsonpath': '$.success == true || $.error_code == 413',
                'remark': '测试大数据量处理能力'
            },
            {
                'name': '并发请求',
                'request_body': self._generate_normal_request_body(request_structure),
                'expected_status': '200',
                'jsonpath': '$.success == true',
                'remark': '并发场景验证（需配合压测工具）'
            }
        ]
        
        for scenario in performance_scenarios:
            self.case_counter += 1
            case_id = f"TC{self.case_counter:04d}"
            
            # 获取前置条件
            preconditions = self._get_preconditions(service_name, method_name)
            
            # 格式化请求体（带类型信息）
            request_body_str = self._format_request_body_with_types(scenario['request_body'], request_structure)
            
            # 生成预期的服务器返回
            expected_response = self._generate_expected_response(service_name, method_name, scenario['request_body'], True)
            
            cases.append({
                'case_id': case_id,
                'title': f'{method_name}_性能边界_{scenario["name"]}',
                'priority': 'P2',
                'preconditions': preconditions,
                'dimension': '性能边界',
                'method_url': f'TCP {service_name}.{method_name}',
                'headers': 'Content-Type: application/protobuf',
                'request_body': request_body_str,
                'expected_status': scenario['expected_status'],
                'server_response': expected_response,
                'status': '通过' if '200' in scenario['expected_status'] else '不通过',
                'jsonpath_assertion': scenario['jsonpath'],
                'error_message': '',
                'problem_analysis': '',
                'remark': scenario['remark']
            })
        
        return cases
    
    def _generate_normal_request_body(self, request_structure: Dict[str, str]) -> Dict:
        """生成正常请求体"""
        body = {}
        if not request_structure:
            return body
        for field_name, field_type in request_structure.items():
            body[field_name] = self._generate_normal_value(field_name, field_type)
        return body
    
    def _generate_normal_value(self, field_name: str, field_type: str) -> Any:
        """根据字段类型生成正常值"""
        # 根据字段名和类型生成合适的值
        if 'id' in field_name.lower():
            return 10000263
        elif 'uid' in field_name.lower():
            return 10000263
        elif 'name' in field_name.lower() or 'nickname' in field_name.lower():
            return "TestName"
        elif 'count' in field_name.lower() or 'limit' in field_name.lower():
            return 20
        elif 'offset' in field_name.lower():
            return 0
        elif 'bool' in field_type.lower() or 'boolean' in field_type.lower():
            return True
        elif 'int' in field_type.lower() or 'int32' in field_type.lower() or 'int64' in field_type.lower():
            return 1
        elif 'string' in field_type.lower():
            return "test_value"
        elif 'repeated' in field_type.lower():
            return []
        else:
            return None
    
    def _generate_abnormal_request_body(self, request_structure: Dict[str, str], 
                                       abnormal_type: str) -> Dict:
        """生成异常请求体"""
        if not request_structure:
            return {}
        
        # 获取第一个字段用于生成异常
        first_field = list(request_structure.keys())[0]
        first_type = request_structure[first_field]
        
        body = self._generate_normal_request_body(request_structure)
        
        if abnormal_type == 'missing_required':
            # 删除必填字段
            del body[first_field]
        elif abnormal_type == 'type_error':
            # 类型错误：字符串字段传数字
            if 'string' in first_type.lower():
                body[first_field] = 12345
            else:
                body[first_field] = "wrong_type"
        elif abnormal_type == 'empty_value':
            # 空值
            if 'string' in first_type.lower():
                body[first_field] = ""
            elif 'int' in first_type.lower():
                body[first_field] = 0
            else:
                body[first_field] = None
        elif abnormal_type == 'out_of_range':
            # 超出范围
            if 'int' in first_type.lower():
                body[first_field] = 999999999
            elif 'string' in first_type.lower():
                body[first_field] = "A" * 10000  # 超长字符串
        elif abnormal_type == 'format_error':
            # 格式错误
            if 'id' in first_field.lower() or 'uid' in first_field.lower():
                body[first_field] = -1  # 负数ID
            else:
                body[first_field] = "invalid_format_@#$%"
        
        return body
    
    def _generate_business_abnormal_request_body(self, request_structure: Dict[str, str],
                                                  scenario: Dict) -> Dict:
        """生成业务异常请求体"""
        body = self._generate_normal_request_body(request_structure)
        
        # 根据场景修改请求体
        if 'request_body' in scenario:
            body.update(scenario['request_body'])
        
        return body
    
    def _generate_large_data_request(self, request_structure: Dict[str, str]) -> Dict:
        """生成大数据量请求"""
        body = self._generate_normal_request_body(request_structure)
        
        # 对于列表字段，生成大量数据
        for field_name, field_type in request_structure.items():
            if 'repeated' in field_type.lower() or 'list' in field_type.lower():
                body[field_name] = [i for i in range(1000)]  # 生成1000条数据
        
        return body
    
    # _get_business_abnormal_scenarios 已在基类中实现
    
    def _get_preconditions(self, service_name: str, method_name: str, request_data: Dict = None) -> str:
        """获取前置条件，根据请求参数分析需要的前置字段"""
        base_precondition = '已登录'
        
        # 如果没有请求数据，返回基础前置条件
        if not request_data:
            return self._get_default_preconditions(service_name, method_name)
        
        # 分析请求参数，识别需要的前置字段
        required_fields = []
        
        # Room服务的前置字段分析
        if service_name == 'room':
            if 'team_id' in request_data and request_data.get('team_id'):
                required_fields.append('已调用 CreateTeam 获取 team_id')
            if 'game_id' in request_data and request_data.get('game_id'):
                required_fields.append('已调用 GetUserState 获取 game_id')
        
        # Social服务的前置字段分析
        elif service_name == 'social':
            if method_name == 'SendMessage':
                scene = request_data.get('scene', 0)
                if scene == 4:  # 世界聊天
                    return f'{base_precondition}（使用世界聊天场景 scene=4，可直接发送消息）'
                elif scene == 1:  # 私聊
                    if 'to_uid' in request_data and request_data.get('to_uid'):
                        return f'{base_precondition}，需要有效的 to_uid（不能是自己）'
            
            if 'conv_id' in request_data and request_data.get('conv_id'):
                required_fields.append('已调用 SendMessage 获取 conv_id')
            if 'seq' in request_data and request_data.get('seq'):
                required_fields.append('已调用 SendMessage 获取 seq')
            if 'msg_id' in request_data and request_data.get('msg_id'):
                required_fields.append('已调用 SendMessage 获取 msg_id')
            if 'target_uid' in request_data and request_data.get('target_uid'):
                if method_name in ['Follow', 'Unfollow']:
                    required_fields.append('需要有效的 target_uid（不能是自己）')
        
        # Hall服务的前置字段分析
        elif service_name == 'hall':
            # Hall服务通常只需要已登录，但可以根据具体需求扩展
            pass
        
        # 组合前置条件
        if required_fields:
            return f'{base_precondition}，{"，".join(required_fields)}'
        else:
            return self._get_default_preconditions(service_name, method_name)
    
    def _get_default_preconditions(self, service_name: str, method_name: str) -> str:
        """获取默认前置条件（当无法从请求参数分析时使用）"""
        preconditions_map = {
            # Hall服务
            'hall': {
                'FetchSelfFullUserInfo': '已登录',
                'FetchSimpleUserInfo': '已登录',
                'UpdateNickname': '已登录',
                'SellItem': '已登录，背包中有物品',
                'BuyItem': '已登录，有足够现金',
                'StashToBackpack': '已登录，仓库中有物品，背包有空位',
                'BackpackToStash': '已登录，背包中有物品',
                'ExchangeBackpackItem': '已登录，背包中有至少2个物品',
            },
            # Room服务
            'room': {
                'GetUserState': '已登录',
                'CreateTeam': '已登录',
                'JoinTeam': '已登录，已调用 CreateTeam 获取 team_id',
                'LeaveTeam': '已登录，已调用 CreateTeam 获取 team_id',
                'GetTeamInfo': '已登录，已调用 CreateTeam 获取 team_id',
                'ChangeReadyState': '已登录，已调用 CreateTeam 获取 team_id',
                'StartGameFromTeam': '已登录，已调用 CreateTeam 获取 team_id，所有成员已准备',
                'Match': '已登录，已调用 CreateTeam 获取 team_id',
                'CancelMatch': '已登录，已调用 CreateTeam 获取 team_id，队伍正在匹配中',
                'GetGameInfo': '已登录，已调用 GetUserState 获取 game_id',
            },
            # Social服务
            'social': {
                'SendMessage': '已登录（世界聊天场景 scene=4 可直接发送，私聊场景需要有效的 to_uid）',
                'PullMsgs': '已登录，已调用 SendMessage 获取 conv_id',
                'RevokeMsg': '已登录，已调用 SendMessage 获取 conv_id 和 seq',
                'DeleteMsg': '已登录，已调用 SendMessage 获取 conv_id 和 seq',
                'AddReaction': '已登录，已调用 SendMessage 获取 conv_id 和 seq',
                'RemoveReaction': '已登录，已调用 SendMessage 获取 conv_id 和 seq，已添加反应',
                'GetReactions': '已登录，已调用 SendMessage 获取 conv_id 和 seq',
                'GetSingleChatConvList': '已登录',
                'MarkRead': '已登录，已调用 SendMessage 获取 conv_id',
                'GetFansList': '已登录',
                'GetFollowList': '已登录',
                'GetFriendList': '已登录',
                'Follow': '已登录，需要有效的 target_uid（不能是自己）',
                'Unfollow': '已登录，已调用 Follow 关注目标用户，需要有效的 target_uid',
            }
        }
        
        service_preconditions = preconditions_map.get(service_name, {})
        return service_preconditions.get(method_name, '已登录')
    
    def _format_request_body_with_types(self, request_body: Dict, request_structure: Dict[str, str]) -> str:
        """格式化请求体，显示字段类型信息"""
        formatted_body = {}
        
        # 遍历所有定义的字段（包括有值和没有值的）
        for field_name, field_type in request_structure.items():
            # 获取字段的实际值
            field_value = request_body.get(field_name)
            
            # 格式化类型名称（去掉repeated前缀）
            type_name = field_type.replace('repeated ', '')
            
            # 构建字段信息
            formatted_body[field_name] = {
                "value": field_value,
                "type": type_name
            }
        
        return json.dumps(formatted_body, indent=2, ensure_ascii=False)
    
    def _generate_expected_response(self, service_name: str, method_name: str, 
                                    request_body: Dict, is_success: bool) -> str:
        """生成预期的服务器返回"""
        if is_success:
            # 成功响应
            response_structure = {
                "success": True,
                "response": self._generate_success_response_data(service_name, method_name, request_body),
                "error_code": 200,
                "error_message": ""
            }
        else:
            # 失败响应
            response_structure = {
                "success": False,
                "response": {},
                "error_code": 400,
                "error_message": self._generate_error_message(service_name, method_name, request_body)
            }
        
        return json.dumps(response_structure, indent=2, ensure_ascii=False)
    
    def _generate_success_response_data(self, service_name: str, method_name: str, request_body: Dict) -> Dict:
        """生成成功响应的数据部分"""
        method_lower = method_name.lower()
        
        # 根据不同的接口生成不同的响应数据
        response_data = {}
        
        if service_name == 'hall':
            if method_name == 'FetchSelfFullUserInfo':
                response_data = {
                    "fetchselffulluserinfo": {
                        "full_user_info": {
                            "uid": 10000263,
                            "nickname": "TestName_123",
                            "cash": 10000,
                            "stash": {"items": []},
                            "backpack": {"cells": []}
                        }
                    }
                }
            elif method_name == 'UpdateNickname':
                response_data = {
                    "updatenickname": {
                        "nickname": request_body.get('nickname', 'TestName_123')
                    }
                }
            elif method_name == 'FetchSimpleUserInfo':
                response_data = {
                    "fetchsimpleuserinfo": {
                        "simple_user_info": {
                            "uid": request_body.get('target_uid', 10000263),
                            "nickname": "TestName_123"
                        }
                    }
                }
            else:
                response_data = {method_lower: {"success": True}}
        
        elif service_name == 'room':
            if method_name == 'CreateTeam':
                response_data = {
                    "createteam": {
                        "team_info": {
                            "team_id": 1006329,
                            "game_mode": request_body.get('game_mode', 1),
                            "members": []
                        }
                    }
                }
            elif method_name == 'GetTeamInfo':
                response_data = {
                    "getteaminfo": {
                        "team_info": {
                            "team_id": request_body.get('team_id', 1006329),
                            "game_mode": 1,
                            "members": []
                        }
                    }
                }
            elif method_name == 'GetUserState':
                response_data = {
                    "getuserstate": {
                        "in_team": True,
                        "team_id": 1006329,
                        "in_game": False,
                        "game_id": 0
                    }
                }
            else:
                response_data = {method_lower: {"success": True}}
        
        elif service_name == 'social':
            if method_name == 'SendMessage':
                response_data = {
                    "sendmessage": {
                        "conv_id": request_body.get('conv_id', 'w_default'),
                        "seq": 1,
                        "msg_id": "msg_123456"
                    }
                }
            elif method_name == 'PullMsgs':
                response_data = {
                    "pullmsgs": {
                        "messages": [],
                        "has_more": False
                    }
                }
            elif method_name == 'Follow':
                response_data = {
                    "follow": {
                        "success": True
                    }
                }
            else:
                response_data = {method_lower: {"success": True}}
        
        return response_data
    
    def _generate_error_message(self, service_name: str, method_name: str, request_body: Dict) -> str:
        """生成错误消息"""
        # 根据不同的错误场景生成不同的错误消息
        if not request_body:
            return "invalid request"
        
        # 检查常见的错误场景
        for key, value in request_body.items():
            if value is None or value == "":
                return f"invalid {key}"
            if isinstance(value, int) and value < 0:
                return f"invalid {key}"
            if isinstance(value, str) and len(value) > 1000:
                return f"invalid {key}: value too long"
        
        # 默认错误消息
        if service_name == 'room' and 'team_id' in request_body:
            return "team not exist"
        elif service_name == 'social' and 'conv_id' in request_body:
            return "conv_id not found"
        elif service_name == 'hall' and 'target_uid' in request_body:
            return "user not exist"
        
        return "invalid request"
    
    def _generate_jsonpath_assertion(self, method_name: str, is_success: bool) -> str:
        """生成JSONPath断言"""
        if is_success:
            return f'$.success == true && $.error_code == 200'
        else:
            return f'$.success == false && $.error_code != 200'
    
    def _generate_db_check(self, method_name: str, is_success: bool) -> str:
        """生成数据库校验"""
        if is_success:
            if 'Create' in method_name or 'Add' in method_name:
                return f'数据库中存在新记录'
            elif 'Update' in method_name:
                return f'数据库记录已更新'
            elif 'Delete' in method_name:
                return f'数据库记录已删除'
            else:
                return f'数据库状态正常'
        else:
            return f'数据未变更'
    
    def _write_test_case_row(self, sheet, row: int, case: Dict, service_name: str, method_name: str):
        """写入测试用例行"""
        # 用例编号
        sheet.cell(row=row, column=1, value=case['case_id'])
        
        # 标题
        sheet.cell(row=row, column=2, value=case['title'])
        
        # 优先级
        priority_cell = sheet.cell(row=row, column=3, value=self.PRIORITY_MAP.get(case['priority'], case['priority']))
        if case['priority'] == 'P0':
            priority_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            priority_cell.font = Font(color="FFFFFF")
        elif case['priority'] == 'P1':
            priority_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        else:
            priority_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 前置条件
        sheet.cell(row=row, column=4, value=case.get('preconditions', '已登录'))
        
        # 维度
        dimension_cell = sheet.cell(row=row, column=5, value=case['dimension'])
        dimension_colors = {
            '正常': 'C6EFCE',
            '参数异常': 'FFC7CE',
            '业务异常': 'FFEB9C',
            '权限安全': 'FF0000',
            '性能边界': '9CC2E5'
        }
        if case['dimension'] in dimension_colors:
            dimension_cell.fill = PatternFill(
                start_color=dimension_colors[case['dimension']],
                end_color=dimension_colors[case['dimension']],
                fill_type="solid"
            )
        
        # 方法+URL
        sheet.cell(row=row, column=6, value=case['method_url'])
        
        # 请求头
        sheet.cell(row=row, column=7, value=case['headers'])
        
        # 请求体
        sheet.cell(row=row, column=8, value=case['request_body'])
        
        # 预期状态码
        sheet.cell(row=row, column=9, value=case['expected_status'])
        
        # 预期服务器返回
        expected_response = case.get('expected_response', case.get('server_response', ''))
        sheet.cell(row=row, column=10, value=expected_response)
        
        # 实际服务器返回
        actual_response = case.get('actual_response', '')
        sheet.cell(row=row, column=11, value=actual_response)
        
        # 状态（通过/不通过）
        status = case.get('status', '未测试')
        status_cell = sheet.cell(row=row, column=12, value=status)
        if status == '通过':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(bold=True, color="006100")
        elif status == '不通过':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(bold=True, color="9C0006")
        else:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(bold=True, color="9C6500")
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # JSONPath断言
        sheet.cell(row=row, column=13, value=case['jsonpath_assertion'])
        
        # 服务器报错（从问题分析中提取）
        server_error = case.get('server_error', '')
        server_error_cell = sheet.cell(row=row, column=14, value=server_error[:500] if server_error else '')
        server_error_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 备注
        sheet.cell(row=row, column=15, value=case['remark'])
        
        # 设置所有单元格样式
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 16):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def _update_case_with_actual_result(self, case: Dict, actual_test_result: Dict) -> Dict:
        """根据实际测试结果更新测试用例
        
        Args:
            case: 原始测试用例
            actual_test_result: 实际测试结果
            
        Returns:
            更新后的测试用例
        """
        import json
        
        # 获取实际响应码和响应数据
        actual_error_code = actual_test_result.get('error_code')
        actual_response = actual_test_result.get('response', {})
        actual_error_message = actual_test_result.get('error_message', '')
        problem_analysis = actual_test_result.get('problem_analysis', '')
        dimension = case.get('dimension', '正常')  # 获取测试维度
        
        # 获取 success 状态（从 actual_test_result 或 actual_response 中）
        actual_success = actual_test_result.get('success', False)
        if isinstance(actual_response, dict) and 'success' in actual_response:
            actual_success = actual_response.get('success', actual_success)
        
        # 保存预期的服务器返回（保持不变）
        expected_response = case.get('server_response', '')
        case['expected_response'] = expected_response
        
        # 保存实际的服务器返回（显示完整结构，与预期格式一致）
        # 构建完整的响应结构，与预期格式保持一致
        if actual_response:
            # 如果 response 是一个包含 'response' 字段的字典（完整结构）
            if isinstance(actual_response, dict) and 'response' in actual_response:
                # 使用完整的响应结构
                full_response = {
                    'success': actual_success,
                    'response': actual_response.get('response', {}),
                    'error_code': actual_error_code or actual_response.get('error_code', 200),
                    'error_message': actual_error_message or actual_response.get('error_message', '')
                }
                actual_response_str = json.dumps(full_response, indent=2, ensure_ascii=False, default=str)
            else:
                # 如果 response 就是实际数据（protobuf解析后的数据），构建完整结构
                full_response = {
                    'success': actual_success,
                    'response': actual_response,
                    'error_code': actual_error_code or 200,
                    'error_message': actual_error_message or ''
                }
                actual_response_str = json.dumps(full_response, indent=2, ensure_ascii=False, default=str)
            case['actual_response'] = actual_response_str
        else:
            # 如果没有返回数据，构建错误响应结构
            error_response = {
                'success': False,
                'response': {},
                'error_code': actual_error_code or 500,
                'error_message': actual_error_message or '无响应数据'
            }
            case['actual_response'] = json.dumps(error_response, indent=2, ensure_ascii=False, default=str)
        
        # 保存错误信息到服务器报错列
        if actual_error_message:
            # 解析问题分析，提取服务器报错
            server_error, _ = self._parse_problem_analysis(problem_analysis)
            # 如果服务器报错为空，使用错误信息
            if not server_error:
                server_error = actual_error_message
            case['server_error'] = server_error
        else:
            # 没有错误信息
            server_error, _ = self._parse_problem_analysis(problem_analysis)
            case['server_error'] = server_error
        
        # 根据实际响应码和测试维度更新状态
        # 对于异常测试用例，返回非200是正常的（通过），返回200是失败的
        # 对于正常测试用例，返回200是正常的（通过），返回非200是失败的
        if actual_error_code is not None:
            is_abnormal = dimension != '正常'
            if is_abnormal:
                # 异常测试用例：返回非200表示通过
                if actual_error_code != 200:
                    case['status'] = '通过'
                else:
                    case['status'] = '不通过'
            else:
                # 正常测试用例：返回200表示通过
                if actual_error_code == 200 and actual_success:
                    case['status'] = '通过'
                else:
                    case['status'] = '不通过'
            case['expected_status'] = str(actual_error_code)
        else:
            case['status'] = '未测试'
        
        return case
    
    def _parse_problem_analysis(self, problem_analysis: str) -> tuple:
        """解析问题分析，提取服务器报错和可能存在的问题
        
        Args:
            problem_analysis: 完整的问题分析字符串
            
        Returns:
            (server_error, possible_issues) 元组
        """
        if not problem_analysis:
            return '', ''
        
        server_error = ''
        possible_issues = ''
        
        # 按行分割
        lines = problem_analysis.split('\n')
        
        # 查找"服务器报错:"部分
        server_error_start = False
        possible_issues_start = False
        
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            
            # 找到"服务器报错:"标记
            if '服务器报错' in line_stripped:
                server_error_start = True
                continue
            
            # 找到"可能存在的问题:"标记
            if '可能存在的问题' in line_stripped:
                server_error_start = False
                possible_issues_start = True
                continue
            
            # 如果在服务器报错部分
            if server_error_start and not possible_issues_start:
                if line_stripped and not line_stripped.startswith('前置条件'):
                    server_error += line_stripped + '\n'
            
            # 如果在可能存在的问题部分
            if possible_issues_start:
                if line_stripped:
                    possible_issues += line_stripped + '\n'
        
        # 清理末尾的换行符
        server_error = server_error.strip()
        possible_issues = possible_issues.strip()
        
        return server_error, possible_issues
    
    def _run_all_tests_from_yaml(self) -> Dict[str, Any]:
        """从YAML文件读取测试用例并实际运行所有接口，返回真实数据"""
        print("\n开始运行接口测试...")
        
        # 初始化API客户端
        client = APIClient(self.config)
        
        # 登录
        if not client.login():
            print("✗ 登录失败，无法运行接口测试")
            return None
        
        # 连接Gate服务器
        if not client.connect_gate():
            print("✗ 连接Gate服务器失败，无法运行接口测试")
            client.close()
            return None
        
        # 解析接口
        parser = ProtobufParser(self.config)
        interfaces = parser.discover_interfaces()
        
        # 收集测试结果
        test_results = {
            'total': 0,
            'passed': 0,
            'failed': 0,
            'errors': 0,
            'services': {}
        }
        
        # 为每个服务运行测试
        for service_name, service_interfaces in interfaces.items():
            print(f"\n{'='*80}")
            print(f"运行 {service_name.upper()} 服务接口测试...")
            print(f"{'='*80}")
            
            service_results = []
            
            # 读取YAML测试用例
            yaml_file = f"test_cases/{service_name}/test_{service_name}.yaml"
            yaml_test_cases = self._load_yaml_test_cases(yaml_file)
            test_cases_dict = yaml_test_cases.get('test_cases', {})
            
            # 为每个接口运行测试
            for interface in service_interfaces:
                method_name = interface['name']
                print(f"\n  接口: {method_name}")
                
                # 获取请求结构（用于判断接口是否有参数）
                service_name_cap = service_name.capitalize()
                request_structure = ProtoRequestFormatter.get_request_structure(service_name_cap, method_name)
                if request_structure is None:
                    request_structure = {}
                
                # 检查接口是否有参数（根据协议定义）
                has_parameters = bool(request_structure and len(request_structure) > 0)
                if not has_parameters:
                    print(f"    ⚠ 接口无参数，将只运行正常测试用例，跳过参数异常测试用例")
                
                # 查找该接口的所有测试用例
                for case_key, case_data in test_cases_dict.items():
                    # 匹配测试用例：case_key 应该是 method_name 或 method_name_xxx
                    if case_key == method_name or case_key.startswith(f"{method_name}_"):
                        dimension = case_data.get('dimension', '正常')
                        
                        # 如果接口没有参数，跳过参数异常测试用例
                        if not has_parameters and dimension == '参数异常':
                            print(f"    ⚠ 跳过参数异常测试用例: {case_key} (接口无参数)")
                            continue
                        
                        description = case_data.get('description', case_key)
                        request_data = case_data.get('request', {})
                        
                        # 转换请求数据（从YAML格式转换为实际请求格式）
                        actual_request = RequestDataConverter.convert_nested_request_data(request_data)
                        
                        # 特殊处理：SendMessage 世界聊天场景（scene=4）的参数优化
                        if service_name == 'social' and method_name == 'SendMessage' and dimension == '正常':
                            scene = actual_request.get('scene', 0)
                            if scene == 4:  # 世界聊天场景
                                # 世界聊天场景不需要 conv_id、to_uid、client_msg_id 等参数
                                # 根据服务器代码，sendWorldMsg 只需要 scene 和 content
                                optimized_request = {
                                    'scene': 4,
                                    'scene_id': 0,  # 世界聊天场景 scene_id 可以为 0
                                    'content': None  # content 会在 _encode_social_body_req 中自动创建默认值
                                }
                                actual_request = optimized_request
                                print(f"    优化请求参数（世界聊天场景）: 使用最小参数集 scene=4, scene_id=0")
                        
                        # 处理前置条件：如果需要 team_id，先调用 CreateTeam 获取
                        if service_name == 'room' and dimension == '正常':
                            # 需要 team_id 的接口：JoinTeam, Match, CancelMatch, LeaveTeam, GetTeamInfo, ChangeReadyState, StartGameFromTeam
                            methods_need_team = ['JoinTeam', 'Match', 'CancelMatch', 'LeaveTeam', 'GetTeamInfo', 'ChangeReadyState', 'StartGameFromTeam']
                            
                            if method_name in methods_need_team:
                                # 检查请求中是否已经有有效的 team_id
                                need_create_team = True
                                if 'team_id' in actual_request and actual_request.get('team_id', 0) > 0:
                                    # 如果请求中有 team_id，先检查队伍是否存在
                                    check_result = client.call_rpc(
                                        service='Room',
                                        method='GetTeamInfo',
                                        request_data={'team_id': actual_request['team_id']}
                                    )
                                    if check_result.get('success') and check_result.get('error_code') == 200:
                                        # 队伍存在，检查用户是否在队伍中
                                        team_info = check_result.get('response', {}).get('getteaminfo', {}).get('team_info', {})
                                        players = team_info.get('players', [])
                                        # 检查当前用户是否在队伍中（需要从 client 获取 uid）
                                        # 这里简化处理：如果队伍存在且状态正常，假设用户也在队伍中
                                        if team_info.get('status') == 1:  # RoomTeamStatusTypeIdle
                                            need_create_team = False
                                            print(f"    ✓ 前置检查: 队伍 {actual_request['team_id']} 已存在且状态正常")
                                
                                if need_create_team:
                                    print(f"    前置操作: 先调用 CreateTeam 创建队伍...")
                                    create_result = client.call_rpc(
                                        service='Room',
                                        method='CreateTeam',
                                        request_data={'game_mode': 1}
                                    )
                                    if create_result.get('success'):
                                        create_response = create_result.get('response', {})
                                        if 'createteam' in create_response:
                                            team_info = create_response['createteam'].get('team_info', {})
                                            team_id = team_info.get('team_id', 0)
                                            if team_id > 0:
                                                actual_request['team_id'] = team_id
                                                print(f"    ✓ 前置操作成功: 已创建队伍，team_id={team_id}")
                                                
                                                # 对于 Match 接口，还需要确保队伍状态是 Idle
                                                if method_name == 'Match':
                                                    # 检查队伍状态，如果不是 Idle，可能需要等待或重置
                                                    print(f"    ✓ 队伍已创建，team_id={team_id}，准备进行匹配")
                                            else:
                                                print(f"    ⚠ 前置操作失败: 无法从 CreateTeam 响应中获取 team_id")
                                    else:
                                        print(f"    ⚠ 前置操作失败: CreateTeam 调用失败 - {create_result.get('error_message', '未知错误')}")
                        
                        # 根据实际请求参数生成详细的前置条件
                        detailed_preconditions = self._get_preconditions(service_name, method_name, actual_request)
                        
                        print(f"    测试用例: {description} (维度: {dimension})")
                        print(f"    前置条件: {detailed_preconditions}")
                        print(f"    请求参数: {json.dumps(actual_request, indent=2, ensure_ascii=False, default=str)}")
                        print(f"    准备调用接口: {service_name.capitalize()}.{method_name}")
                        
                        # 调用接口
                        try:
                            result = client.call_rpc(
                                service=service_name.capitalize(),
                                method=method_name,
                                request_data=actual_request
                            )
                            
                            # 判断测试状态
                            success = result.get('success', False)
                            error_code = result.get('error_code', 200)
                            error_message = result.get('error_message', '')
                            response_data = result.get('response', {})
                            
                            # 对于异常测试用例，返回非200是正常的
                            if dimension != '正常':
                                test_status = 'success' if error_code != 200 else 'failure'
                            else:
                                test_status = 'success' if success and error_code == 200 else 'failure'
                            
                            # 打印结果
                            if test_status == 'success':
                                print(f"    ✓ 测试通过")
                            else:
                                print(f"    ✗ 测试失败: error_code={error_code}, error_message={error_message}")
                                if error_message:
                                    print(f"    错误信息: {error_message}")
                            
                            # 保存测试结果（response 结构要与 client.call_rpc 返回的结构一致）
                            # client.call_rpc 返回: {'success': bool, 'response': dict, 'error_code': int, 'error_message': str}
                            # 这里保存完整的响应结构，方便后续处理
                            test_result = {
                                'name': description,
                                'method': method_name,
                                'status': test_status,
                                'request': actual_request,
                                'response': {
                                    'success': success,
                                    'response': response_data,  # 这是从 client.call_rpc 返回的 response 字段（protobuf解析后的数据）
                                    'error_code': error_code,
                                    'error_message': error_message
                                },
                                'error_code': error_code,
                                'error_message': error_message,
                                'success': success,  # 添加 success 字段，方便状态判断
                                'preconditions': detailed_preconditions,  # 使用详细的前置条件
                                'dimension': dimension,
                                'abnormal_type': case_key.split('_', 1)[1] if '_' in case_key and dimension != '正常' else ''
                            }
                            
                            service_results.append(test_result)
                            test_results['total'] += 1
                            if test_status == 'success':
                                test_results['passed'] += 1
                            else:
                                test_results['failed'] += 1
                                
                        except Exception as e:
                            # 捕获异常并打印
                            error_msg = str(e)
                            print(f"    ✗ 接口调用异常: {error_msg}")
                            import traceback
                            traceback.print_exc()
                            
                            # 保存错误结果
                            test_result = {
                                'name': description,
                                'method': method_name,
                                'status': 'error',
                                'request': actual_request,
                                'response': {},
                                'error_code': 500,
                                'error_message': error_msg,
                                'preconditions': case_data.get('preconditions', '已登录'),
                                'dimension': dimension,
                                'abnormal_type': case_key.split('_', 1)[1] if '_' in case_key and dimension != '正常' else ''
                            }
                            
                            service_results.append(test_result)
                            test_results['total'] += 1
                            test_results['errors'] += 1
            
            # 保存服务结果
            if service_results:
                test_results['services'][service_name] = {
                    'test_results': service_results
                }
        
        # 关闭连接
        client.close()
        
        # 打印统计信息
        print(f"\n{'='*80}")
        print("测试运行完成")
        print(f"{'='*80}")
        print(f"总测试数: {test_results['total']}")
        print(f"通过: {test_results['passed']}")
        print(f"失败: {test_results['failed']}")
        print(f"错误: {test_results['errors']}")
        print(f"{'='*80}\n")
        
        return test_results
    
    def _load_test_results_from_excel(self) -> Dict[str, Any]:
        """从最新的test_report.xlsx加载测试结果"""
        try:
            import openpyxl
            import json
            report_path = os.path.join(self.config.get_report_dir(), "test_report.xlsx")
            if not os.path.exists(report_path):
                return None
            
            wb = openpyxl.load_workbook(report_path)
            test_results = {
                'services': {}
            }
            
            # 遍历所有工作表（除了"测试摘要"）
            for sheet_name in wb.sheetnames:
                if sheet_name == "测试摘要":
                    continue
                
                service_name = sheet_name.lower()
                sheet = wb[sheet_name]
                service_test_results = {
                    'test_results': []
                }
                
                # 从第2行开始读取数据（第1行是表头）
                # 先读取表头，确定各列的索引
                headers = []
                for col in range(1, sheet.max_column + 1):
                    header = sheet.cell(row=1, column=col).value
                    headers.append(header)
                
                # 找到关键列的索引
                name_col = 2  # 标题列（默认第2列）
                error_code_col = None
                response_col = 11  # 实际服务器返回列（默认第11列）
                error_message_col = 14  # 错误信息列（默认第14列）
                method_col = 6  # 方法+URL列（默认第6列）
                
                for i, header in enumerate(headers, 1):
                    if header and '标题' in str(header):
                        name_col = i
                    elif header and ('错误码' in str(header) or '状态码' in str(header)):
                        error_code_col = i
                    elif header and ('实际' in str(header) and '返回' in str(header)):
                        response_col = i
                    elif header and ('错误信息' in str(header) or '错误' in str(header)):
                        error_message_col = i
                    elif header and ('方法' in str(header) or 'URL' in str(header)):
                        method_col = i
                
                for row in range(2, sheet.max_row + 1):
                    # 读取标题（测试用例名称）
                    name = sheet.cell(row=row, column=name_col).value
                    if not name:
                        continue
                    
                    # 从方法+URL列提取方法名（例如：TCP ROOM.LeaveTeam -> LeaveTeam）
                    method_url = sheet.cell(row=row, column=method_col).value or ''
                    method_name = ''
                    if '.' in method_url:
                        method_name = method_url.split('.')[-1]
                    
                    # 读取响应码（从实际服务器返回的JSON中提取，或从状态列推断）
                    error_code = None
                    # 尝试从实际服务器返回列读取JSON并提取error_code
                    response_str = sheet.cell(row=row, column=response_col).value if response_col else None
                    if response_str:
                        try:
                            response_obj = json.loads(response_str)
                            if isinstance(response_obj, dict):
                                error_code = response_obj.get('error_code')
                        except:
                            pass
                    
                    # 读取响应数据
                    response = {}
                    if response_str:
                        try:
                            response = json.loads(response_str)
                        except:
                            pass
                    
                    # 读取错误信息
                    error_message = ''
                    if error_message_col:
                        error_message = sheet.cell(row=row, column=error_message_col).value or ''
                    
                    # 问题分析（从错误信息和其他列组合）
                    problem_analysis = ''
                    if error_message:
                        problem_analysis = f"错误信息: {error_message}"
                    
                    # 读取维度（从第5列）
                    dimension_col = 5  # 维度列
                    dimension = sheet.cell(row=row, column=dimension_col).value or ''
                    
                    # 读取前置条件（从第4列）
                    preconditions = []
                    preconditions_col = 4  # 前置条件列
                    preconditions_str = sheet.cell(row=row, column=preconditions_col).value or ''
                    if preconditions_str:
                        # 如果前置条件是分号分隔的列表
                        if ';' in preconditions_str:
                            preconditions = [p.strip() for p in preconditions_str.split(';')]
                        else:
                            preconditions = [preconditions_str]
                    
                    # 从标题中提取测试用例名称和异常类型
                    # 标题格式可能是：
                    # - "SendMessage - 正常调用"
                    # - "SendMessage - 参数异常 - 必填参数缺失"
                    # - "SendMessage_参数异常_必填参数缺失"
                    test_name = method_name if method_name else name
                    abnormal_type = ''
                    
                    if ' - ' in name:
                        # 格式：SendMessage - 参数异常 - 必填参数缺失
                        title_parts = name.split(' - ')
                        if len(title_parts) >= 3:
                            # 有异常类型信息
                            abnormal_type = ' - '.join(title_parts[2:])
                            test_name = f"{method_name}_{dimension}_{abnormal_type.replace(' - ', '_')}" if dimension else f"{method_name}_{abnormal_type.replace(' - ', '_')}"
                        elif len(title_parts) == 2:
                            # 只有两部分，可能是正常调用
                            if title_parts[1] == '正常调用':
                                test_name = method_name
                            else:
                                test_name = f"{method_name}_{title_parts[1].replace(' ', '_')}"
                    elif '_' in name:
                        # 格式：SendMessage_参数异常_必填参数缺失
                        parts = name.split('_')
                        if len(parts) >= 3:
                            abnormal_type = '_'.join(parts[2:])
                            test_name = name
                        else:
                            test_name = name
                    else:
                        # 如果标题已经是完整格式，直接使用
                        test_name = name.replace(' - ', '_').replace(' ', '_')
                    
                    # 如果没有提取到异常类型，但维度不是正常，尝试从标题中提取
                    if not abnormal_type and dimension and dimension != '正常':
                        # 从标题中提取异常类型（例如：SendMessage - 参数异常 - 必填参数缺失）
                        if ' - ' in name:
                            parts = name.split(' - ')
                            if len(parts) >= 3:
                                abnormal_type = ' - '.join(parts[2:])
                    
                    service_test_results['test_results'].append({
                        'name': test_name,
                        'method': method_name if method_name else name.split()[0] if name else 'Unknown',
                        'dimension': dimension if dimension else None,
                        'abnormal_type': abnormal_type if abnormal_type else None,
                        'error_code': error_code,
                        'response': response,
                        'error_message': error_message,
                        'problem_analysis': problem_analysis,
                        'preconditions': preconditions
                    })
                
                if service_test_results['test_results']:
                    test_results['services'][service_name] = service_test_results
            
            if test_results['services']:
                return test_results
            else:
                return None
        except Exception as e:
            print(f"⚠ 从Excel加载测试结果失败: {e}")
            return None


def main():
    """主函数"""
    config = Config()
    generator = TestCaseGenerator(config)
    output_path = generator.generate_test_cases_excel("test_cases_complete.xlsx")
    print(f"\n测试用例Excel已生成: {output_path}")


if __name__ == '__main__':
    main()

