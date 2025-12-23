"""
HTML报告生成器 - 生成测试报告
"""
import os
import json
from datetime import datetime
from typing import Dict, Any
from jinja2 import Template
from framework.config import Config

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠ openpyxl 未安装，Excel导出功能将不可用。请运行: pip install openpyxl")



class ReportGenerator:
    """HTML报告生成器"""
    
    def __init__(self, config: Config):
        self.config = config
        self.report_dir = config.get_report_dir()
        os.makedirs(self.report_dir, exist_ok=True)
    
    def generate_report(self, test_results: Dict[str, Any]) -> str:
        """生成HTML和Excel报告"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"test_report_{timestamp}"
        report_path = os.path.join(self.report_dir, f"{report_filename}.html")
        
        # 生成HTML报告
        html_content = self._generate_html(test_results)
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        # 生成Excel报告
        excel_path = self._generate_excel(test_results, "test_report.xlsx")
        
        # 清理旧报告
        self._cleanup_old_reports()
        
        return report_path
    
    def _generate_html(self, test_results: Dict[str, Any]) -> str:
        """生成HTML内容"""
        template_str = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>API测试报告</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
            border-bottom: 3px solid #4CAF50;
            padding-bottom: 10px;
        }
        .summary {
            display: flex;
            gap: 20px;
            margin: 30px 0;
        }
        .stat-card {
            flex: 1;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            color: white;
        }
        .stat-card.total { background: #2196F3; }
        .stat-card.passed { background: #4CAF50; }
        .stat-card.failed { background: #f44336; }
        .stat-number {
            font-size: 36px;
            font-weight: bold;
            margin: 10px 0;
        }
        .service-section {
            margin: 40px 0;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            overflow: hidden;
        }
        .service-header {
            background: #f8f9fa;
            padding: 15px 20px;
            font-size: 18px;
            font-weight: bold;
            color: #333;
            border-bottom: 2px solid #e0e0e0;
            cursor: pointer;
        }
        .service-header:hover {
            background: #e9ecef;
        }
        .service-content {
            padding: 20px;
            display: none;
        }
        .service-content.expanded {
            display: block;
        }
        .test-case {
            background: #f8f9fa;
            margin: 10px 0;
            padding: 15px;
            border-radius: 4px;
            border-left: 4px solid #4CAF50;
        }
        .test-case.failed {
            border-left-color: #f44336;
        }
        .test-case.info {
            border-left-color: #FFA500;
        }
        .test-name {
            font-weight: bold;
            margin-bottom: 10px;
        }
        .test-detail {
            margin: 5px 0;
            font-size: 14px;
        }
        .detail-label {
            font-weight: bold;
            color: #666;
        }
        pre {
            background: #f5f5f5;
            padding: 10px;
            border-radius: 4px;
            overflow-x: auto;
            font-size: 12px;
        }
        .problem-analysis {
            background: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 10px;
            margin-top: 10px;
        }
        .problem-analysis .detail-label {
            color: #856404;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>API测试报告</h1>
        <p>生成时间: {{ timestamp }}</p>
        
        <div class="summary">
            <div class="stat-card total">
                <div class="stat-label">总接口数</div>
                <div class="stat-number">{{ test_results.total }}</div>
            </div>
            <div class="stat-card passed">
                <div class="stat-label">通过</div>
                <div class="stat-number">{{ test_results.passed }}</div>
            </div>
            <div class="stat-card failed">
                <div class="stat-label">失败</div>
                <div class="stat-number">{{ test_results.failed }}</div>
            </div>
        </div>
        
        {% for service_name, service_data in test_results.services.items() %}
        <div class="service-section">
            <div class="service-header" onclick="toggleService('{{ service_name }}')">
                {{ service_name|upper }} 服务
            </div>
            <div class="service-content" id="service-{{ service_name }}">
                {% if service_data.test_results|length == 0 %}
                <div class="test-case info">
                    <div class="test-name">暂无正常测试用例</div>
                    <div class="test-detail">
                        <span class="detail-label">提示:</span>
                        <pre>该服务暂无正常测试用例（只有异常测试用例，异常测试用例在Excel报告中查看）</pre>
                    </div>
                </div>
                {% else %}
                {% for test in service_data.test_results %}
                <div class="test-case {{ 'failed' if test.status == 'failure' else ('info' if test.status == 'info' else '') }}">
                    <div class="test-name">{{ test.name }}</div>
                    {% if test.status != 'info' %}
                    <div class="test-detail">
                        <span class="detail-label">请求:</span>
                        <pre>{{ test.request }}</pre>
                    </div>
                    <div class="test-detail">
                        <span class="detail-label">响应:</span>
                        <pre>{{ test.response }}</pre>
                    </div>
                    {% endif %}
                    {% if test.error_message %}
                    <div class="test-detail">
                        <span class="detail-label">{% if test.status == 'info' %}提示:{% else %}错误:{% endif %}</span>
                        <pre>{{ test.error_message }}</pre>
                    </div>
                    {% endif %}
                    {% if test.error %}
                    <div class="test-detail">
                        <span class="detail-label">错误:</span>
                        <pre>{{ test.error }}</pre>
                    </div>
                    {% endif %}
                    {% if test.problem_analysis %}
                    <div class="test-detail problem-analysis">
                        <span class="detail-label">问题分析:</span>
                        <pre>{{ test.problem_analysis }}</pre>
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
                {% endif %}
            </div>
        </div>
        {% endfor %}
    </div>
    
    <script>
        function toggleService(serviceName) {
            const content = document.getElementById('service-' + serviceName);
            content.classList.toggle('expanded');
        }
    </script>
</body>
</html>'''
        
        template = Template(template_str)
        
        # 准备数据 - 将字典转换为JSON字符串
        def format_json(obj):
            # 如果是请求参数（包含value和type的结构），格式化显示
            if isinstance(obj, dict) and obj and isinstance(list(obj.values())[0], dict):
                # 检查是否是格式化的请求结构（包含value和type）
                first_value = list(obj.values())[0]
                if 'value' in first_value and 'type' in first_value:
                    # 格式化请求参数，显示字段名、类型和值
                    formatted = {}
                    for field_name, field_info in obj.items():
                        if isinstance(field_info, dict) and 'value' in field_info and 'type' in field_info:
                            value = field_info['value']
                            field_type = field_info['type']
                            if value is not None:
                                formatted[f"{field_name} ({field_type})"] = value
                            else:
                                formatted[f"{field_name} ({field_type})"] = "[未提供]"
                        else:
                            formatted[field_name] = field_info
                    obj = formatted
            
            # 递归转换 protobuf 对象为字典
            def convert_to_dict(val):
                # 检查是否是 protobuf 消息对象
                if hasattr(val, 'DESCRIPTOR') and hasattr(val, 'SerializeToString'):
                    try:
                        from google.protobuf.json_format import MessageToDict
                        return MessageToDict(val, including_default_value_fields=True, preserving_proto_field_name=True)
                    except:
                        # 如果 MessageToDict 失败，尝试手动转换
                        result = {}
                        try:
                            for field_descriptor in val.DESCRIPTOR.fields:
                                field_name = field_descriptor.name
                                field_value = getattr(val, field_name)
                                if field_descriptor.label == field_descriptor.LABEL_REPEATED:
                                    # 重复字段，需要递归转换每个元素
                                    if field_descriptor.type == field_descriptor.TYPE_MESSAGE:
                                        result[field_name] = [convert_to_dict(item) for item in field_value]
                                    else:
                                        result[field_name] = list(field_value)
                                elif field_descriptor.type == field_descriptor.TYPE_MESSAGE:
                                    # 嵌套消息
                                    if field_value:
                                        result[field_name] = convert_to_dict(field_value)
                                else:
                                    # 基本类型
                                    result[field_name] = field_value
                        except Exception as e:
                            # 如果转换失败，返回字符串表示
                            return str(val)
                        return result
                elif isinstance(val, dict):
                    return {k: convert_to_dict(v) for k, v in val.items()}
                elif isinstance(val, list):
                    return [convert_to_dict(item) for item in val]
                else:
                    return val
            
            try:
                converted = convert_to_dict(obj)
                return json.dumps(converted, indent=2, ensure_ascii=False, default=str)
            except Exception as e:
                # 如果转换失败，返回字符串表示
                return str(obj)
        
        # 处理测试结果，将字典转换为JSON字符串
        # 先统计正常测试用例的数量（过滤掉异常测试用例）
        normal_test_count = 0
        normal_passed_count = 0
        normal_failed_count = 0
        normal_errors_count = 0
        
        # 加载所有YAML测试用例，用于更新标题
        yaml_test_cases_map = {}
        for service_name in test_results.get('services', {}).keys():
            yaml_file = f"test_cases/{service_name}/test_{service_name}.yaml"
            if os.path.exists(yaml_file):
                try:
                    import yaml
                    with open(yaml_file, 'r', encoding='utf-8') as f:
                        yaml_data = yaml.safe_load(f) or {}
                        yaml_test_cases_map[service_name] = yaml_data.get('test_cases', {})
                except Exception as e:
                    print(f"⚠ 加载YAML测试用例失败 {yaml_file}: {e}")
                    yaml_test_cases_map[service_name] = {}
            else:
                yaml_test_cases_map[service_name] = {}
        
        for service_name, service_data in test_results.get('services', {}).items():
            yaml_test_cases = yaml_test_cases_map.get(service_name, {})
            for test in service_data.get('test_results', []):
                # 从YAML中查找对应的测试用例，更新标题
                method_name = test.get('method', test.get('name', 'Unknown'))
                test_name = test.get('name', method_name)
                
                # 尝试匹配YAML中的测试用例
                yaml_case_key = None
                yaml_case_data = {}
                for key in yaml_test_cases.keys():
                    if key.startswith(method_name) or key == method_name or key == f"{method_name}_正常":
                        yaml_case_key = key
                        yaml_case_data = yaml_test_cases[key]
                        break
                
                # 如果没找到，尝试根据dimension和abnormal_type匹配
                if not yaml_case_data:
                    dimension = test.get('dimension', '正常')
                    abnormal_type = test.get('abnormal_type', '')
                    if dimension != '正常' and abnormal_type:
                        for key in yaml_test_cases.keys():
                            if key.startswith(method_name) and abnormal_type in key:
                                yaml_case_key = key
                                yaml_case_data = yaml_test_cases[key]
                                break
                        if not yaml_case_data:
                            abnormal_keywords = abnormal_type.split('_') if '_' in abnormal_type else [abnormal_type]
                            for key in yaml_test_cases.keys():
                                if key.startswith(method_name):
                                    if any(keyword in key for keyword in abnormal_keywords if keyword):
                                        yaml_case_key = key
                                        yaml_case_data = yaml_test_cases[key]
                                        break
                
                # 如果找到YAML数据，使用description作为标题
                if yaml_case_data and yaml_case_data.get('description'):
                    test['name'] = yaml_case_data.get('description')
                
                # 过滤异常测试用例：只统计正常测试用例
                dimension = test.get('dimension')
                abnormal_type = test.get('abnormal_type')
                
                # 如果有dimension且不是'正常'，或者有abnormal_type，则跳过
                if dimension and dimension != '正常':
                    continue
                if abnormal_type:
                    continue
                
                # 统计正常测试用例
                normal_test_count += 1
                status = test.get('status', 'unknown')
                if status == 'success':
                    normal_passed_count += 1
                elif status == 'failure':
                    normal_failed_count += 1
                elif status == 'error':
                    normal_errors_count += 1
        
        processed_results = {
            'total': normal_test_count,
            'passed': normal_passed_count,
            'failed': normal_failed_count,
            'errors': normal_errors_count,
            'services': {}
        }
        
        for service_name, service_data in test_results.get('services', {}).items():
            processed_results['services'][service_name] = {
                'test_results': []
            }
            normal_count = 0
            yaml_test_cases = yaml_test_cases_map.get(service_name, {})
            for test in service_data.get('test_results', []):
                # 过滤异常测试用例：只显示正常测试用例
                dimension = test.get('dimension')
                abnormal_type = test.get('abnormal_type')
                
                # 如果有dimension且不是'正常'，或者有abnormal_type，则跳过（不显示在HTML中）
                if dimension and dimension != '正常':
                    continue
                if abnormal_type:
                    continue
                
                # 从YAML中查找对应的测试用例，更新标题
                method_name = test.get('method', test.get('name', 'Unknown'))
                test_name = test.get('name', method_name)
                
                # 尝试匹配YAML中的测试用例
                yaml_case_data = {}
                for key in yaml_test_cases.keys():
                    if key.startswith(method_name) or key == method_name or key == f"{method_name}_正常":
                        yaml_case_data = yaml_test_cases[key]
                        break
                
                # 如果找到YAML数据，使用description作为标题
                if yaml_case_data and yaml_case_data.get('description'):
                    test_name = yaml_case_data.get('description')
                
                normal_count += 1
                processed_test = {
                    'name': test_name,  # 使用从YAML获取的description作为标题
                    'status': test.get('status', 'unknown'),
                    'request': format_json(test.get('request', {})),
                    'response': format_json(test.get('response', {})),
                    'error': test.get('error', ''),
                    'error_code': test.get('error_code', ''),
                    'error_message': test.get('error_message', ''),
                    'problem_analysis': test.get('problem_analysis', '')
                }
                processed_results['services'][service_name]['test_results'].append(processed_test)
            
            # 如果没有正常测试用例，添加提示信息
            if normal_count == 0:
                processed_results['services'][service_name]['test_results'].append({
                    'name': '提示',
                    'status': 'info',
                    'request': '',
                    'response': '',
                    'error': '',
                    'error_code': '',
                    'error_message': '该服务暂无正常测试用例（只有异常测试用例，异常测试用例在Excel报告中查看）',
                    'problem_analysis': ''
                })
        
        html_data = {
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'test_results': processed_results
        }
        
        return template.render(**html_data)
    
    def _generate_excel(self, test_results: Dict[str, Any], filename: str) -> str:
        """生成Excel报告（按照指定格式）"""
        if not OPENPYXL_AVAILABLE:
            print("⚠ Excel导出功能不可用，跳过Excel报告生成")
            return ""
        
        excel_path = os.path.join(self.report_dir, filename)
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认工作表
        
        # 为每个服务创建工作表
        for service_name, service_data in test_results.get('services', {}).items():
            sheet = wb.create_sheet(title=service_name.upper())
            self._write_service_excel_sheet(sheet, service_name, service_data)
        
        # 保存文件
        wb.save(excel_path)
        print(f"✓ Excel报告已生成: {excel_path}")
        
        return excel_path
    
    def _write_service_excel_sheet(self, sheet, service_name: str, service_data: Dict[str, Any]):
        """写入服务Excel工作表（直接使用测试运行的实际结果，不进行推断）"""
        # 不再从YAML加载数据，直接使用测试运行的实际结果
        
        # 样式定义
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        failure_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头（14列格式，删除"服务器返回"列）
        headers = [
            '用例编号', '标题', '优先级', '前置条件', '维度',
            '方法+URL', '请求头', '请求', '预期状态码', '预期',
            '实际服务器返回', '状态', 'JSONPath断言', '错误信息'
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # 写入测试数据（包括正常测试用例和异常测试用例）
        row = 2
        case_counter = 1
        
        # 遍历所有测试结果，直接使用测试运行的实际结果，不进行任何推断或从YAML填写
        test_results_list = service_data.get('test_results', [])
        print(f"📊 {service_name.upper()} 服务: 准备写入 {len(test_results_list)} 个测试结果到Excel")
        if len(test_results_list) == 0:
            print(f"⚠️  警告: {service_name.upper()} 服务没有测试结果！")
        else:
            # 打印前5个测试用例的名称，用于调试
            for i, test in enumerate(test_results_list[:5]):
                test_name = test.get('name', test.get('method', 'Unknown'))
                print(f"  - 测试用例 {i+1}: {test_name}")
        
        for test in test_results_list:
            # 直接使用测试运行的实际结果，不进行任何推断
            method_name = test.get('method', test.get('name', 'Unknown'))
            test_name = test.get('name', method_name)
            dimension = test.get('dimension', '正常')
            abnormal_type = test.get('abnormal_type', '')
            
            # 用例编号
            case_id = f"TC{case_counter:04d}"
            sheet.cell(row=row, column=1, value=case_id)
            
            # 标题（优先使用 test_name，如果包含异常类型信息则使用，否则根据 dimension 和 abnormal_type 构建）
            title = test_name
            # 如果test_name是测试方法名格式（test_开头），去掉前缀
            if title.startswith('test_'):
                title = title[5:]  # 去掉test_前缀
            
            # 检查 title 是否已经包含异常类型信息（包含下划线或"异常"等关键词）
            has_abnormal_info = '_' in title or '异常' in title or '安全' in title or '性能' in title or '边界' in title
            
            if has_abnormal_info:
                # 异常测试用例，将下划线转换为" - "
                title = title.replace('_', ' - ')
            elif dimension != '正常' and abnormal_type:
                # 如果 dimension 不是正常，且有 abnormal_type，构建完整标题
                title = f"{method_name} - {dimension} - {abnormal_type}"
            elif dimension != '正常':
                # 如果只有 dimension，构建标题
                title = f"{method_name} - {dimension}"
            else:
                # 正常测试用例，添加"正常调用"后缀
                title = f"{title} - 正常调用"
            
            sheet.cell(row=row, column=2, value=title)
            
            # 优先级（默认P1，不从YAML推断）
            priority = 'P1'
            priority_cell = sheet.cell(row=row, column=3, value=priority)
            priority_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            
            # 前置条件（直接使用test_result中的preconditions，如果没有则使用默认值）
            preconditions = test.get('preconditions', [])
            if isinstance(preconditions, list):
                preconditions_str = '; '.join(preconditions) if preconditions else '已登录'
            else:
                preconditions_str = str(preconditions) if preconditions else '已登录'
            sheet.cell(row=row, column=4, value=preconditions_str)
            
            # 维度（直接使用test_result中的dimension）
            dimension = test.get('dimension', '正常')
            dimension_cell = sheet.cell(row=row, column=5, value=dimension)
            dimension_colors = {
                '正常': 'C6EFCE',
                '参数异常': 'FFC7CE',
                '业务异常': 'FFEB9C',
                '权限安全': 'FF0000',
                '性能边界': '9CC2E5'
            }
            if dimension in dimension_colors:
                dimension_cell.fill = PatternFill(
                    start_color=dimension_colors[dimension],
                    end_color=dimension_colors[dimension],
                    fill_type="solid"
                )
            
            # 方法+URL（TCP协议）
            request_method = test.get('request_method', 'TCP')
            method_url = f"{request_method} {service_name.upper()}.{method_name}"
            sheet.cell(row=row, column=6, value=method_url)
            
            # 请求头（TCP协议使用protobuf）
            headers_str = 'Content-Type: application/protobuf'
            sheet.cell(row=row, column=7, value=headers_str)
            
            # 请求（直接使用实际运行的请求数据）
            request_data = test.get('request', {})
            # 如果request_data是包含value和type的格式，提取实际值
            if isinstance(request_data, dict):
                formatted_request = {}
                for key, value in request_data.items():
                    if isinstance(value, dict) and 'value' in value:
                        formatted_request[key] = value.get('value')
                    else:
                        formatted_request[key] = value
                request_data = formatted_request
            
            try:
                request_str = json.dumps(request_data, indent=2, ensure_ascii=False, default=str) if request_data else '{}'
            except:
                request_str = str(request_data) if request_data else '{}'
            sheet.cell(row=row, column=8, value=request_str[:5000])
            
            # 预期状态码（使用默认值200，不从YAML推断）
            expected_status = '200'
            sheet.cell(row=row, column=9, value=expected_status)
            
            # 预期（空，不从YAML推断）
            expected_str = ''
            sheet.cell(row=row, column=10, value=expected_str)
            
            # 实际服务器返回（直接使用实际运行的响应数据，不进行任何推断或修改）
            response_data = test.get('response', {})
            # 直接使用response_data，不进行任何推断或修改
            actual_response_data = response_data if response_data else {}
            
            try:
                actual_response_str = json.dumps(actual_response_data, indent=2, ensure_ascii=False, default=str) if actual_response_data else ''
            except:
                actual_response_str = str(actual_response_data) if actual_response_data else ''
            sheet.cell(row=row, column=11, value=actual_response_str[:10000])
            
            # 状态（根据实际服务器返回的响应码判断）
            # 对于异常测试用例，返回非200错误码是正常的（通过），返回200是失败的
            # 对于正常测试用例，返回200是正常的（通过），返回非200是失败的
            if isinstance(response_data, dict) and 'error_code' in response_data:
                actual_error_code = response_data.get('error_code')
            else:
                actual_error_code = test.get('error_code')
            
            # 判断是否为异常测试用例
            is_abnormal_test = (dimension and dimension != '正常') or test.get('abnormal_type')
            
            if is_abnormal_test:
                # 异常测试用例：返回非200错误码表示通过，返回200表示失败
                if actual_error_code and actual_error_code != 200:
                    status_text = '通过'
                    status_cell = sheet.cell(row=row, column=12, value=status_text)
                    status_cell.fill = success_fill
                else:
                    status_text = '失败'
                    status_cell = sheet.cell(row=row, column=12, value=status_text)
                    status_cell.fill = failure_fill
            else:
                # 正常测试用例：返回200表示通过，返回非200表示失败
                if actual_error_code == 200:
                    status_text = '通过'
                    status_cell = sheet.cell(row=row, column=12, value=status_text)
                    status_cell.fill = success_fill
                else:
                    status_text = '失败'
                    status_cell = sheet.cell(row=row, column=12, value=status_text)
                    status_cell.fill = failure_fill
            
            # JSONPath断言（根据测试类型使用默认断言，不从YAML推断）
            dimension = test.get('dimension', '正常')
            if dimension and dimension != '正常':
                jsonpath_assertion = '$.error_code != 200'  # 异常测试用例默认断言
            else:
                jsonpath_assertion = '$.success == true && $.error_code == 200'  # 正常测试用例默认断言
            sheet.cell(row=row, column=13, value=jsonpath_assertion)
            
            # 错误信息（从test_result）
            error_msg = test.get('error_message', '')
            if not error_msg:
                error_msg = test.get('error', '')
            if not error_msg and isinstance(response_data, dict):
                error_msg = response_data.get('error_message', '')
            
            # 提取关键错误信息
            if error_msg:
                error_lines = error_msg.split('\n')
                key_error_lines = []
                for line in error_lines:
                    line = line.strip()
                    if line and ('AssertionError' in line or 'False is not true' in line or 'API调用失败' in line or '异常测试失败' in line):
                        key_error_lines.append(line)
                if key_error_lines:
                    error_msg = '\n'.join(key_error_lines)
                if len(error_msg) > 2000:
                    error_msg = error_msg[:1000] + '\n...\n' + error_msg[-1000:]
            
            sheet.cell(row=row, column=14, value=error_msg[:2000] if error_msg else '')
            
            # 应用样式
            for col in range(1, 15):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            row += 1
            case_counter += 1
        
        # 调整列宽
        sheet.column_dimensions['A'].width = 12  # 用例编号
        sheet.column_dimensions['B'].width = 30  # 标题
        sheet.column_dimensions['C'].width = 10  # 优先级
        sheet.column_dimensions['D'].width = 30  # 前置条件
        sheet.column_dimensions['E'].width = 12  # 维度
        sheet.column_dimensions['F'].width = 25  # 方法+URL
        sheet.column_dimensions['G'].width = 30  # 请求头
        sheet.column_dimensions['H'].width = 50  # 请求
        sheet.column_dimensions['I'].width = 15  # 预期状态码
        sheet.column_dimensions['J'].width = 50  # 预期
        sheet.column_dimensions['K'].width = 60  # 实际服务器返回
        sheet.column_dimensions['L'].width = 10  # 状态
        sheet.column_dimensions['M'].width = 30  # JSONPath断言
        sheet.column_dimensions['N'].width = 50  # 错误信息
        
        # 设置行高
        for row_idx in range(2, row):
            sheet.row_dimensions[row_idx].height = 60
    
    def _cleanup_old_reports(self, keep_count: int = 3):
        """清理旧报告"""
        try:
            report_files = []
            if os.path.exists(self.report_dir):
                for filename in os.listdir(self.report_dir):
                    if filename.startswith('test_report_') and filename.endswith('.html'):
                        filepath = os.path.join(self.report_dir, filename)
                        if os.path.isfile(filepath):
                            mtime = os.path.getmtime(filepath)
                            report_files.append((mtime, filepath, filename))
            
            report_files.sort(key=lambda x: x[0], reverse=True)
            
            if len(report_files) > keep_count:
                for mtime, filepath, filename in report_files[keep_count:]:
                    os.remove(filepath)
                    print(f"删除旧报告: {filename}")
        except Exception as e:
            print(f"⚠ 清理旧报告失败: {e}")

